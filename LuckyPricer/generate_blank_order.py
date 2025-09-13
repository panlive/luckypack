"""
generate_blank_order.py — Генерация Excel "БЛАНК ЗАКАЗА"

-----------------------------------------------------
Назначение:
    • Находит файл data/selected.json (создан через select_products.py)
    • Генерирует Excel-файл для клиента в data/exports/
    • Соблюдает формат LuckyPack, всё оформлено

Логирование:
    • Все сообщения — в logs/blank_order.log
    • Ошибки дублируются админу через admin_notify

-----------------------------------------------------
"""

import pandas as pd
import os
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import logging
from config import EXPORTS_DIR, LOGS_DIR
from admin_notify import notify_admin

LOG_FILE = os.path.join(LOGS_DIR, "blank_order.log")
logging.basicConfig(
    filename=LOG_FILE,
    filemode='a',
    format='%(asctime)s [%(levelname)s] %(message)s',
    level=logging.INFO,
    encoding='utf-8'
)

SELECTED_JSON_PATH = "data/selected.json"

def log(msg, level='info'):
    print(msg)
    getattr(logging, level)(msg)

def generate_excel(client_name="тестового клиента"):
    try:
        df = pd.read_json(SELECTED_JSON_PATH)

        now = pd.Timestamp.now(tz="Europe/Moscow")
        date_str = now.strftime("%d.%m.%Y %H-%M МСК")

        filename = f"Подбор товаров для {client_name} — {date_str}.xlsx"
        filepath = os.path.join(EXPORTS_DIR, filename)

        df["Артикул"] = df["Артикул"].astype(str)

        df.columns = [
            "Артикул",
            "Номенклатура, Характеристика, Упаковка",
            "Шт/Кор",
            "Опт с НДС",
            "Опт с НДС от 150 000 руб.",
            "Спец Цена",
        ]
        df["Ваш заказ"] = ""

        if os.path.exists(filepath):
            os.remove(filepath)

        df.to_excel(filepath, index=False, startrow=1)

        wb = load_workbook(filepath)
        ws = wb.active

        ws.cell(row=1, column=1).value = f"Подбор товаров для «{client_name}» — {now.strftime('%d.%m.%Y %H:%M')} МСК"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
        ws.cell(row=1, column=1).font = Font(size=14, bold=True)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

        for col in range(1, 8):
            cell = ws.cell(row=2, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
            for cell in row[:2]:
                cell.alignment = Alignment(horizontal="left")
            for cell in row[2:]:
                cell.alignment = Alignment(horizontal="right")

        thin = Side(border_style="thin", color="000000")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=7):
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=6, max_col=6):
            for cell in row:
                if cell.value not in [None, "", 0]:
                    cell.font = Font(color="FF0000")

        ws.cell(row=2, column=6).font = Font(color="FF0000", bold=True)

        fill = PatternFill(fill_type="solid", start_color="FFFACD", end_color="FFFACD")
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=7, max_col=7):
            for cell in row:
                cell.fill = fill

        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

        wb.save(filepath)
        log(f"✅ Бланк заказа создан: {filepath}")

    except Exception as e:
        err = f"❌ Ошибка при создании бланка заказа: {e}"
        log(err, 'error')
        notify_admin(err, module="generate_blank_order.py")

# Аргумент из командной строки
if __name__ == "__main__":
    import sys
    client_name = sys.argv[1] if len(sys.argv) > 1 else "тестового клиента"
    generate_excel(client_name)