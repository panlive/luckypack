#!/usr/bin/env python3
import os, re, json, argparse, asyncio, datetime
from pathlib import Path
from dotenv import load_dotenv
from aiogram import Bot
from aiogram.types import InputMediaPhoto, InputFile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

PROJ = Path("/srv/luckypack/App")
DATA = Path("/app/data/photos")
OUT  = Path("/app/data/PhotoPicks"); OUT.mkdir(parents=True, exist_ok=True)

load_dotenv(PROJ/".env")
BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
CHAT_ID   = int(os.getenv("SUPERADMIN_ID","0"))

# --- цветовые токены ---
RED_STRICT = ["красн"]  # строго "красный"
RED_SHADES = [
  "красн","бордо","алый","рубинов","вишн","кармин","корал","терракот",
  "марсал","малин","гранат","пурпур","фиолетов-красн","розов","фукс",
]

def load_products():
    p = PROJ/"LuckyPricer"/"products.json"
    data = json.load(open(p,"r",encoding="utf-8"))
    # список словарей
    return data if isinstance(data,list) else []

def is_film(rec:str)->bool:
    cat  = str(rec.get("Категория","")).lower()
    name = str(rec.get("Наименование") or rec.get("Номенклатура, Характеристика, Упаковка") or "").lower()
    return ("пленк" in cat) or ("пленк" in name)

def match_color(name:str, mode:str)->bool:
    s = (name or "").lower()
    toks = RED_STRICT if mode=="strict" else RED_SHADES
    return any(t in s for t in toks)

def find_photo(art:str):
    t = DATA/"thumbs"/f"{art}.webp"
    v = DATA/"vectorized"/f"{art}.webp"
    if t.exists(): return str(t)
    if v.exists(): return str(v)
    return None

# --- XLSX (без фото — как базовый бланк) ---
BOLD=Font(b=True); CENTER=Alignment(horizontal="center",vertical="center",wrap_text=True)
LEFT=Alignment(horizontal="left",vertical="top",wrap_text=True)
RIGHT=Alignment(horizontal="right",vertical="center")
GRAY=PatternFill("solid", fgColor="F2F2F2")
from openpyxl.styles import Side, Border
THIN=Side(style="thin", color="DDDDDD"); BORDER=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)

COLS=[("Артикул",16),("Наименование",60),("ШТ/КОР",10),("ОПТ с НДС",14),("ОПТ с НДС от 150 000 руб.",22),("СПЕЦ ЦЕНА",14),("Ваш заказ",12)]

def build_xlsx(rows, title, path):
    wb=Workbook(); ws=wb.active; ws.title="Подбор по запросу"
    head=f"Подбор по запросу — {title} — {datetime.datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(COLS))
    c=ws.cell(row=1,column=1,value=head); c.font=Font(b=True,size=12); c.alignment=CENTER
    for j,(name,w) in enumerate(COLS, start=1):
        cell=ws.cell(row=2,column=j,value=name); cell.font=BOLD; cell.alignment=CENTER; cell.fill=GRAY; cell.border=BORDER
        ws.column_dimensions[cell.column_letter].width=w
    ws.freeze_panes="A3"; ws.auto_filter.ref=f"A2:{chr(64+len(COLS))}2"
    r=3
    for rec in rows:
        ws.cell(row=r,column=1,value=rec.get("Артикул","")).alignment=CENTER
        ws.cell(row=r,column=2,value=rec.get("Наименование","")).alignment=LEFT
        ws.cell(row=r,column=3,value=rec.get("ШТ/КОР","")).alignment=RIGHT
        ws.cell(row=r,column=4,value=rec.get("ОПТ с НДС","")).alignment=RIGHT
        ws.cell(row=r,column=5,value=rec.get("ОПТ с НДС от 150 000 руб.","")).alignment=RIGHT
        ws.cell(row=r,column=6,value=rec.get("СПЕЦ ЦЕНА","")).alignment=RIGHT
        ws.cell(row=r,column=7,value="").alignment=RIGHT
        for j in range(1,len(COLS)+1): ws.cell(row=r,column=j).border=BORDER
        r+=1
    wb.save(path); return path

async def send_demo(query_text:str, mode:str, top:int):
    bot=Bot(BOT_TOKEN, parse_mode=None)

    prods=[r for r in load_products() if is_film(r)]
    for r in prods:
        r["Наименование"]=str(r.get("Наименование") or r.get("Номенклатура, Характеристика, Упаковка") or "")
    matched=[r for r in prods if match_color(r["Наименование"], mode)]
    matched.sort(key=lambda x: (x.get("Категория",""), x.get("Наименование","")))
    total=len(matched)

    # 1) «уточняющий» текст + что делаем в демо
    clar = ("Запрос: «%s».\nПредлагаю два режима: 1) только «красный», 2) все оттенки красного.\n"
            "В демо подобрал по режиму: **все оттенки**.") % query_text
    await bot.send_message(chat_id=CHAT_ID, text=clar)

    # 2) Медиа-группа топ-10 без цен в подписях
    media=[]
    shown=0
    for r in matched[:top]:
        art=str(r.get("Артикул",""))
        name=r.get("Наименование","")
        p=find_photo(art)
        if not p: continue
        cap=f"[{len(media)+1:02d}] Артикул: {art}\n{name}"
        media.append(InputMediaPhoto(media=InputFile(p), caption=cap))
        shown+=1
    if media:
        await bot.send_media_group(chat_id=CHAT_ID, media=media)

    # 3) XLSX с полным списком (без фото)
    title=f"Клиент: Сергей | Запрос: {query_text} ({'все оттенки' if mode!='strict' else 'только красный'})"
    xlsx_path = OUT / f"Подбор по запросу — {title}.xlsx"
    build_xlsx(matched, title, xlsx_path)
    await bot.send_document(chat_id=CHAT_ID, document=InputFile(str(xlsx_path)),
                            caption=f"Всего найдено: {total}. В ленте показано: {shown}. Полный список — в файле.\nНужно показать ещё в чате — скажи: Ещё.")

    await bot.session.close()

def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("--q", default="Плёнка матовая красная")
    ap.add_argument("--mode", choices=["strict","shades"], default="shades")
    ap.add_argument("--top", type=int, default=10)
    args=ap.parse_args()
    asyncio.run(send_demo(args.q, args.mode, args.top))

if __name__=="__main__":
    main()
