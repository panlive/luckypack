#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from __future__ import annotations
import argparse, datetime, json, sys
from pathlib import Path
from typing import Dict, List, Optional
import numpy as np
from PIL import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

PROJ = Path("/srv/luckypack/project")
DATA = Path("/app/data/photos")
EMB  = Path("/srv/luckypack/project/SearchByPhoto/index")
PROD_JSON = PROJ / "LuckyPricer/products.json"
PROD_JSONS_DIR = PROJ / "LuckyPricer/data/jsons"
PNG_CACHE = Path("/app/data/PhotoPicks/_png"); PNG_CACHE.mkdir(parents=True, exist_ok=True)

# Попытка подключить FAISS (для оффлайн-поиска по уже индексированному артикулу)
try:
    import faiss  # type: ignore
except Exception:
    try:
        import faiss_cpu as faiss  # type: ignore
    except Exception:
        faiss = None  # режим поиска по артикулу будет недоступен

HEADERS = ["Фото","Артикул","Наименование","ШТ/КОР","ОПТ с НДС","ОПТ с НДС от 150 000 руб.","СПЕЦ ЦЕНА","Ваш заказ"]
ALIGN_LEFT      = Alignment(horizontal="left",  vertical="center", wrap_text=True)
ALIGN_LEFT_TOP  = Alignment(horizontal="left",  vertical="top",   wrap_text=True)
ALIGN_CENTER    = Alignment(horizontal="center",vertical="center", wrap_text=True)
ALIGN_RIGHT     = Alignment(horizontal="right", vertical="center")
THIN = Side(style="thin", color="DDDDDD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
YELLOW = PatternFill(fill_type="solid", start_color="FFFAE6", end_color="FFFAE6")  # нежная пастель

def now_msk_str()->str:
    dt = datetime.datetime.utcnow() + datetime.timedelta(hours=3)
    return dt.strftime("%d.%m.%Y %H:%M МСК")

def _load_products_from_products_json(path: Path)->Dict[str,Dict[str,str]]:
    if not path.exists(): return {}
    try:
        arr = json.loads(path.read_text(encoding="utf-8"))
        if not isinstance(arr, list) or not arr: return {}
    except Exception:
        return {}
    out = {}
    for rec in arr:
        art = str(rec.get("Артикул","")).strip()
        if not art: continue
        name = rec.get("Наименование") or rec.get("Номенклатура, Характеристика, Упаковка") or rec.get("Номенклатура") or ""
        out[art] = {
            "Наименование": str(name),
            "ШТ/КОР": str(rec.get("ШТ/КОР","")),
            "ОПТ с НДС": str(rec.get("ОПТ с НДС","")),
            "ОПТ с НДС от 150 000 руб.": str(rec.get("ОПТ с НДС от 150 000 руб.","")),
            "СПЕЦ ЦЕНА": str(rec.get("СПЕЦ ЦЕНА","")),
        }
    return out

def _load_products_from_jsons_dir(dir_path: Path)->Dict[str,Dict[str,str]]:
    out = {}
    if not dir_path.exists(): return out
    for p in sorted(dir_path.glob("*.json")):
        try:
            arr = json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            continue
        if not isinstance(arr, list): continue
        for rec in arr:
            art = str(rec.get("Артикул","")).strip()
            if not art: continue
            name = rec.get("Наименование") or rec.get("Номенклатура, Характеристика, Упаковка") or rec.get("Номенклатура") or ""
            out[art] = {
                "Наименование": str(name),
                "ШТ/КОР": str(rec.get("ШТ/КОР","")),
                "ОПТ с НДС": str(rec.get("ОПТ с НДС","")),
                "ОПТ с НДС от 150 000 руб.": str(rec.get("ОПТ с НДС от 150 000 руб.","")),
                "СПЕЦ ЦЕНА": str(rec.get("СПЕЦ ЦЕНА","")),
            }
    return out

def load_products()->Dict[str,Dict[str,str]]:
    prod = _load_products_from_products_json(PROD_JSON)
    return prod if prod else _load_products_from_jsons_dir(PROD_JSONS_DIR)

def ensure_png_for_excel(article: str)->Optional[Path]:
    p_png = PNG_CACHE / f"{article}.png"
    if p_png.exists(): return p_png
    p_thumb = DATA / "thumbs" / f"{article}.webp"
    p_vect  = DATA / "vectorized" / f"{article}.webp"
    src = p_thumb if p_thumb.exists() else (p_vect if p_vect.exists() else None)
    if src is None: return None
    try:
        Image.open(src).convert("RGBA").save(p_png)
        return p_png
    except Exception:
        return None

def _num(v: str):
    if not v: return None
    try:
        return float(v.replace(" ","").replace("\\u00A0","").replace(",","."))
    except Exception:
        return None

def _style_title(ws, title: str):
    ws.merge_cells("A1:H1")
    c = ws["A1"]; c.value = title; c.font = Font(b=True, size=14)
    c.alignment = ALIGN_LEFT
    ws.row_dimensions[1].height = 28

def _style_contacts(ws):
    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value = "Контакты: [Имя] | Тел.: [Телефон] | Email: [Email] | Город/доставка: [Город] | Менеджер: [Менеджер]"
    c.alignment = ALIGN_LEFT
    ws.row_dimensions[2].height = 20

def _set_headers(ws, row: int = 3):
    ws.row_dimensions[row].height = 24
    for j, title in enumerate(HEADERS, start=1):
        cell = ws.cell(row=row, column=j, value=title)
        cell.font = Font(b=True); cell.alignment = ALIGN_CENTER; cell.border = BORDER
    ws.column_dimensions["A"].width = 15.0  # Фото

def _auto_fit(ws, start_row: int, end_row: int):
    for col_idx, letter in enumerate(["B","C","D","E","F","G","H"], start=2):
        max_len = 0
        for r in range(start_row-1, end_row+1):  # включая заголовок
            v = ws.cell(row=r, column=col_idx).value
            if v is None: continue
            ln = len(str(v))
            if ln > max_len: max_len = ln
        width = min(max(10, int(max_len*1.1) + 2), 90)
        if letter == "H": width = max(width, 12)
        ws.column_dimensions[letter].width = float(width)

def build_excel(arts: List[str], products: Dict[str,Dict[str,str]], title: str, out_path: Path)->Path:
    wb = Workbook(); ws = wb.active
    _style_title(ws, title)
    _style_contacts(ws)
    header_row = 4
    _set_headers(ws, header_row)

    r = header_row + 1  # первая строка данных = 4
    if not arts:
        ws.cell(row=r, column=3, value="совпадений нет").alignment = ALIGN_LEFT_TOP
        r += 1
    else:
        for art in arts:
            rec = products.get(art, {})
            p_png = ensure_png_for_excel(art)
            if p_png and p_png.exists():
                img = XLImage(str(p_png))
                MAX = 90
                ow, oh = getattr(img, "width", None), getattr(img, "height", None)
                if isinstance(ow,(int,float)) and isinstance(oh,(int,float)) and ow>0 and oh>0:
                    k = min(MAX/ow, MAX/oh, 1.0); img.width = int(ow*k); img.height = int(oh*k)
                ws.row_dimensions[r].height = 75
                img.anchor = f"A{r}"; ws.add_image(img)
            ws.cell(row=r, column=2, value=art).alignment = ALIGN_LEFT_TOP
            ws.cell(row=r, column=3, value=rec.get("Наименование","")).alignment = ALIGN_LEFT_TOP
            ws.cell(row=r, column=4, value=rec.get("ШТ/КОР","")).alignment = ALIGN_RIGHT

            v5=_num(rec.get("ОПТ с НДС","")); c5=ws.cell(row=r, column=5, value=v5 if v5 is not None else rec.get("ОПТ с НДС",""))
            if v5 is not None: c5.number_format="#,##0.00"; c5.alignment=ALIGN_RIGHT
            else: c5.alignment=ALIGN_RIGHT

            v6=_num(rec.get("ОПТ с НДС от 150 000 руб.","")); c6=ws.cell(row=r, column=6, value=v6 if v6 is not None else rec.get("ОПТ с НДС от 150 000 руб.",""))
            if v6 is not None: c6.number_format="#,##0.00"; c6.alignment=ALIGN_RIGHT
            else: c6.alignment=ALIGN_RIGHT

            v7=_num(rec.get("СПЕЦ ЦЕНА","")); c7=ws.cell(row=r, column=7, value=v7 if v7 is not None else rec.get("СПЕЦ ЦЕНА",""))
            if v7 is not None: c7.number_format="#,##0.00"; c7.alignment=ALIGN_RIGHT
            else: c7.alignment=ALIGN_RIGHT

            ws.cell(row=r, column=8, value="").alignment = ALIGN_RIGHT
            r += 1

    # границы для всего диапазона до минимум 35-й строки
    last_data_row = max(r-1, 35)
    for rr in range(header_row+1, last_data_row+1):
        for cc in range(1, len(HEADERS)+1):
            cell = ws.cell(row=rr, column=cc)
            if cell.value is None: cell.value = ""  # чтобы границы были видны
            cell.border = BORDER
        ws.cell(row=rr, column=8).fill = YELLOW  # вся колонка H

    # заголовок H тоже жёлтый
    ws["H"+str(header_row)].fill = YELLOW

    _auto_fit(ws, header_row, last_data_row)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return out_path

def _load_img_ids()->List[str]:
    p = EMB / "img_ids.npy"
    if not p.exists(): raise FileNotFoundError(f"Нет файла: {p}")
    arr = np.load(str(p), allow_pickle=True)
    out = []
    for x in arr: out.append(x.decode("utf-8","ignore") if isinstance(x,(bytes,bytearray)) else str(x))
    return out

def _load_faiss_index():
    if faiss is None:
        print("ERROR: FAISS не установлен в этом окружении.", file=sys.stderr)
        return None
    idx_path = EMB / "faiss_img.index"
    if not idx_path.exists():
        print(f"ERROR: Не найден индекс: {idx_path}", file=sys.stderr)
        return None
    try:
        return faiss.read_index(str(idx_path))
    except Exception as e:
        print(f"ERROR: Не удалось прочитать индекс: {e}", file=sys.stderr)
        return None

def _neighbors_by_article(article: str, topk: int = 5, include_self: bool = False) -> List[str]:
    """
    Возвращает список соседей (артикулы) по уже индексированному артикулу.
    Никаких сетевых вызовов. Требует support reconstruct() у индекса.
    """
    ids = _load_img_ids()
    base = Path(article.strip()).stem  # снимаем расширение, если передали имя файла
    lookup = {s: i for i, s in enumerate(ids)}
    if base not in lookup:
        raise KeyError(f"Артикул '{article}' не найден в индексе.")
    row = lookup[base]
    index = _load_faiss_index()
    if index is None:
        raise RuntimeError("FAISS недоступен.")
    try:
        vec = index.reconstruct(row)
    except Exception:
        raise RuntimeError("Этот тип FAISS-индекса не поддерживает reconstruct().")
    if vec is None:
        raise RuntimeError("Не удалось получить вектор для указанного артикула.")
    xq = np.asarray(vec, dtype="float32").reshape(1, -1)
    k = int(topk) + (0 if include_self else 1)
    k = max(1, k)
    D, I = index.search(xq, k)
    res: List[str] = []
    for idx in I[0].tolist():
        if not include_self and idx == row:
            continue
        if 0 <= idx < len(ids):
            res.append(str(ids[idx]))
        if len(res) >= topk:
            break
    return res

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--smoke", action="store_true"); ap.add_argument("--n", type=int, default=10)
    ap.add_argument("--by-article", help="Артикул из индекса для поиска соседей")
    ap.add_argument("--topk", type=int, default=5, help="Сколько соседей вернуть")
    ap.add_argument("--include-self", action="store_true", help="Не убирать сам запрос из выдачи")
    ap.add_argument("--as-json", action="store_true", help="Вывод в JSON (stdout), без текстовых строк")
    ap.add_argument("--excel", required=True); ap.add_argument("--title", default="")
    args = ap.parse_args()
    title = args.title.strip() or f"Подбор по фото — {now_msk_str()}"
    products = load_products()

    if args.by_article:
        try:
            arts = _neighbors_by_article(args.by_article, topk=args.topk, include_self=args.include_self)
        except Exception as e:
            print(f"ERROR: {e}", file=sys.stderr)
            sys.exit(2)
    else:
        ids = _load_img_ids()
        arts = ids[:args.n] if args.smoke else []

    out = build_excel(arts, products, title, Path(args.excel))

    if args.as_json:
        print(json.dumps({"query": (args.by_article or None), "neighbors": arts, "excel": str(out)}, ensure_ascii=False))
    else:
        print(f"OK: XLSX сохранён: {out}")
        if arts:
            print("RESULT:", ", ".join(arts))
        else:
            print("Совпадений не найдено — файл создан с пометкой.")

if __name__ == "__main__":
    main()
